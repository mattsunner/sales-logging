"""
Sales Log Generator

Author: Matthew Sunner, 2020
"""

# Imports
from openpyxl import Workbook, load_workbook
from datetime import date

def salesLogging():
    # Import sample list
    sample_txt = open('sample.txt', 'r')

    # Variables 
    sales_items = []
    today = date.today()
    author = str(input("Enter Your Name: "))

    # Iterate through list and add each item to an array
    for line in sample_txt:
        sales_items.append(line.strip())

    # Iterate through the array and create a new, standardized workbook for each item
    for item in sales_items:
        wb = load_workbook('original.xlsx')
        ws = wb.active
        ws['B3'] = item
        ws['B4'] = today
        ws['B5'] = author
        wb.save(f'workbooks/Sales-Log-{item}.xlsx')
        wb.close()

if __name__ == '__main__':
    salesLogging()